import streamlit as st
import datetime
import calendar
import pandas as pd
import altair as alt
from utils.sections import festivos, Section
import numpy as np
import os
import json
import matplotlib
import io
from datetime import datetime as datetime_type
from navigation import make_sidebar
from utils.worker import Worker
from utils.db import get_db
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Set page config
st.set_page_config(page_title="📅 Calendari de guàrdies", layout="wide")
make_sidebar()
db = get_db()
# Get sections and workers from the database instead of JSON files
sections = db.get_sections()

# Convert the list of fechas in each section to datetime objects if needed
for section in sections:
    if hasattr(section, 'fechas') and section.fechas:
        section.fechas = [fecha if isinstance(fecha, datetime.date) else 
                         datetime_type.strptime(fecha, '%Y-%m-%d').date() 
                         for fecha in section.fechas]

# Get workers from the database
workers = db.get_workers()


def get_day_label(date_obj, festivos_list):
    """Converts a date to day label (monday, tuesday, etc.)"""
    if date_obj in festivos_list:
        return "festivo"
    
    weekday = date_obj.weekday()
    day_map = {
        0: "monday", 1: "tuesday", 2: "wednesday", 3: "thursday",
        4: "friday", 5: "saturday", 6: "sunday",
    }
    return day_map[weekday]

def day_matches_section(date_obj, section_obj, festivos_list):
    """
    Returns True if section applies to the given date.
    Critical logic:
    1. Day of week must match section's days or be marked as festivo
    2. If section has specific fechas (dates), the date must be in that list
       (HEMS and Coordis have specific dates they operate)

    params:
    - date_obj: Date to check
    - section_obj: Section object to check against
    """
    day_label = get_day_label(date_obj, festivos_list)
    
    # First check if day matches section's days
    if day_label not in section_obj.dias:
        return False
    
    # Then check if section has specific dates
    if hasattr(section_obj, 'fechas') and section_obj.fechas:
        return date_obj in section_obj.fechas
        
    # If no specific dates defined, it applies to all days matching the weekday
    return True

def get_shifts_data(year, month=None):
    """Get shifts data for visualization"""
    data = []
    
    # If month is specified, only get data for that month
    if month:
        start_date = datetime.date(year, month, 1)
        if month == 12:
            end_date = datetime.date(year + 1, 1, 1)
        else:
            end_date = datetime.date(year, month + 1, 1)
    else:
        start_date = datetime.date(year, 1, 1)
        end_date = datetime.date(year + 1, 1, 1)
    
    current_date = start_date
    while current_date < end_date:
        for sec in sections:
            if day_matches_section(current_date, sec, festivos):
                data.append({
                    "date": current_date,
                    "day": current_date.day,
                    "month": current_date.month,
                    "weekday": get_day_label(current_date, festivos),
                    "shift_name": sec.nombre,
                    "hours": sec.horas_turno,
                    "personnel": sec.personal,
                    "libra": sec.libra,
                    "is_festivo": current_date in festivos
                })
        current_date += datetime.timedelta(days=1)
    
    return pd.DataFrame(data)

def get_shift_color(shift_name):
    """Return color based on shift category"""
    if "HEMS" in shift_name:
        return "#4285F4"  # Blue
    elif "Coordis" in shift_name:
        return "#34A853"  # Green
    elif "UCI" in shift_name:
        return "#FBBC05"  # Yellow/Orange
    elif "Urg" in shift_name:
        return "#EA4335"  # Red
    else:
        return "#9AA0A6"  # Gray

def draw_month_calendar(df_month, month_name, year):
    """Draw a calendar for a specific month using Streamlit"""
    # Calculate the first day of week for this month and the number of days
    if df_month.empty:
        first_day = datetime.date(year, month_names.index(month_name) + 1, 1)
    else:
        first_day = datetime.date(df_month['date'].min().year, df_month['date'].min().month, 1)
    
    first_weekday = first_day.weekday()  # 0 for Monday, 6 for Sunday
    days_in_month = calendar.monthrange(first_day.year, first_day.month)[1]
    
    # Create a 7x6 grid (7 days x max 6 weeks)
    weeks = 6
    
    # Display day headers (Mon, Tue, etc.)
    cols = st.columns(7)
    weekday_names = ["Dilluns", "Dimarts", "Dimecres", "Dijous", "Divendres", "Dissabte", "Diumenge"]
    for i, col in enumerate(cols):
        col.markdown(f"<div style='text-align: center; font-weight: bold;'>{weekday_names[i]}</div>", unsafe_allow_html=True)
    
    # Create a 2D list to hold all cells
    day_num = 1 - first_weekday
    
    for w in range(weeks):
        cols = st.columns(7)
        for d in range(7):
            current_day = day_num
            day_num += 1
            
            # Skip days outside the month
            if current_day < 1 or current_day > days_in_month:
                cols[d].markdown("<div style='background-color: #eaeaea; border: 1px solid #ddd; padding: 5px; min-height: 80px; border-radius: 5px;'>&nbsp;</div>", unsafe_allow_html=True)
                continue
            
            # Create date object for current day
            day_date = datetime.date(first_day.year, first_day.month, current_day)
            
            # Get shifts for this day
            if df_month.empty:
                day_shifts = pd.DataFrame()
            else:
                day_shifts = df_month[df_month['date'] == day_date]
            
            # Construct cell content with all shifts
            cell_content = f"<div style='text-align: right; font-weight: bold; margin-bottom: 5px;'>{current_day}</div>"
            
            for _, shift in day_shifts.iterrows():
                color = get_shift_color(shift['shift_name'])
                cell_content += f"<div style='background-color: {color}; color: white; padding: 2px 4px; margin: 2px 0; border-radius: 3px; font-size: 0.8em; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;'>{shift['shift_name']}</div>"
            
            # Style for the cell
            is_festivo = day_date in festivos
            if is_festivo:
                cell_style = "background-color: #ffcccc; border: 1px solid #ddd; padding: 5px; min-height: 80px; border-radius: 5px;"
            else:
                cell_style = "background-color: #f9f9f9; border: 1px solid #ddd; padding: 5px; min-height: 80px; border-radius: 5px;"
            
            # Render the cell
            cols[d].markdown(f'<div style="{cell_style}">{cell_content}</div>', unsafe_allow_html=True)
            
        # Break if we've displayed all days in the month
        if day_num > days_in_month:
            break

def draw_assignment_calendar(df_month, assignments_df, month_name, year):
    """Draw a calendar with shift assignments for a specific month"""
    
    # Calculate the first day of week for this month and the number of days
    if df_month.empty:
        first_day = datetime.date(year, month_names.index(month_name) + 1, 1)
    else:
        first_day = datetime.date(df_month['date'].min().year, df_month['date'].min().month, 1)
    
    first_weekday = first_day.weekday()  # 0 for Monday, 6 for Sunday
    days_in_month = calendar.monthrange(first_day.year, first_day.month)[1]
    
    # Create a 7x6 grid (7 days x max 6 weeks)
    weeks = 6
    
    # Display day headers (Mon, Tue, etc.)
    cols = st.columns(7)
    weekday_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for i, col in enumerate(cols):
        col.markdown(f"<div style='text-align: center; font-weight: bold;'>{weekday_names[i]}</div>", unsafe_allow_html=True)
    
    # Create a 2D list to hold all cells
    day_num = 1 - first_weekday
    
    for w in range(weeks):
        cols = st.columns(7)
        for d in range(7):
            current_day = day_num
            day_num += 1
            
            # Skip days outside the month
            if current_day < 1 or current_day > days_in_month:
                cols[d].markdown("<div style='background-color: #eaeaea; border: 1px solid #ddd; padding: 5px; min-height: 80px; border-radius: 5px;'>&nbsp;</div>", unsafe_allow_html=True)
                continue
            
            # Create date object for current day
            day_date = datetime.date(first_day.year, first_day.month, current_day)
            
            # Get shifts for this day
            if df_month.empty:
                day_shifts = pd.DataFrame()
            else:
                day_shifts = df_month[df_month['date'] == day_date]
            
            # Get assignments for this day
            if assignments_df.empty:
                day_assignments = pd.DataFrame()
            else:
                day_assignments = assignments_df[assignments_df['date'].dt.date == day_date]
            
            # Construct cell content with the day number
            cell_content = f"<div style='text-align: right; font-weight: bold; margin-bottom: 5px;'>{current_day}</div>"
            
            # Add shifts with assignments
            for _, shift in day_shifts.iterrows():
                color = get_shift_color(shift['shift_name'])
                
                # Check if this shift has been assigned
                assigned_worker = ""
                matching_assignments = day_assignments[day_assignments['section_name'] == shift['shift_name']]
                if not matching_assignments.empty:
                    worker_name = matching_assignments.iloc[0]['worker_name']
                    assigned_worker = ""
                    for worker in workers:
                        if worker.name == worker_name:
                            assigned_worker = f": {worker.initials}"
                            break
                    
                    # If we couldn't find the worker, create initials from name
                    if not assigned_worker and worker_name:
                        # Create initials from the name (first letter of each word)
                        name_parts = worker_name.split()
                        initials = ''.join([part[0].upper() for part in name_parts if part])[:2]
                        assigned_worker = f": {initials}"
                
                cell_content += f"<div style='background-color: {color}; color: white; padding: 2px 4px; margin: 2px 0; border-radius: 3px; font-size: 0.8em; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;'>{shift['shift_name']}{assigned_worker}</div>"
            
            # Style for the cell
            is_festivo = day_date in festivos
            if is_festivo:
                cell_style = "background-color: #ffcccc; border: 1px solid #ddd; padding: 5px; min-height: 80px; border-radius: 5px;"
            else:
                cell_style = "background-color: #f9f9f9; border: 1px solid #ddd; padding: 5px; min-height: 80px; border-radius: 5px;"
            
            # Render the cell
            cols[d].markdown(f'<div style="{cell_style}">{cell_content}</div>', unsafe_allow_html=True)
            
        # Break if we've displayed all days in the month
        if day_num > days_in_month:
            break


def create_calendar_excel(df_year: pd.DataFrame, ass: bool = False, ass_df : pd.DataFrame | None = None) -> bytes | None:
    """
    Create an in-memory Excel file (as bytes) with one sheet per month.
    Expects columns: ['date', 'weekday', 'shift_name', 'is_festivo'].
    Returns bytes on success, or None if there's no data/sheets.
    """
    required_cols = {"date", "weekday", "shift_name", "is_festivo"}
    missing = required_cols - set(df_year.columns)
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")

    df = df_year[["date", "weekday", "shift_name", "is_festivo"]].copy()
    if df.empty:
        return None

    # Ensure datetime dtype
    df["date"] = pd.to_datetime(df["date"])

    # Month helpers
    df["month_num"] = df["date"].dt.month
    # Use Python's calendar for consistent month names
    df["month_name"] = df["date"].dt.month_name()
    
    # Order rows consistently
    df = df.sort_values(["date", "shift_name"], kind="stable")
    # Create a new DataFrame with only the data we need for the Excel file
    # And replace English weekday names with Spanish
    spanish_weekdays = {
        "monday": "Lunes",
        "tuesday": "Martes",
        "wednesday": "Miércoles",
        "thursday": "Jueves",
        "friday": "Viernes",
        "saturday": "Sábado",
        "sunday": "Domingo",
        "festivo": "Festivo"
    }
    
    # Replace weekdays in the dataframe
    df["weekday"] = df["weekday"].map(lambda day: spanish_weekdays.get(day.lower(), day))
    # Prepare output buffer + writer
    excel_data = io.BytesIO()
    sheets_created = 0

    # Different colors for each weekday and festivo
    monday_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    tuesday_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")  # Light blue
    wednesday_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Light yellow
    thursday_fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")  # Light pink
    friday_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")  # Light olive
    saturday_fill = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")  # Light purple
    sunday_fill = PatternFill(start_color="FAC090", end_color="FAC090", fill_type="solid")  # Light orange
    festivo_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

    # Create a mapping of weekday to fill color
    weekday_fills = {
        0: monday_fill,  # Monday
        1: tuesday_fill,  # Tuesday
        2: wednesday_fill,  # Wednesday
        3: thursday_fill,  # Thursday
        4: friday_fill,  # Friday
        5: saturday_fill,  # Saturday
        6: sunday_fill,  # Sunday
        "festivo": festivo_fill  # Festivo
    }
    header_align = Alignment(horizontal="center", vertical="center")

    with pd.ExcelWriter(excel_data, engine="openpyxl") as writer:
        # Iterate months in calendar order (1..12) and only if present
        for month_num in sorted(df["month_num"].unique()):
            month_df = df[df["month_num"] == month_num].copy()
            if month_df.empty:
                continue

            # Human-friendly columns (keep weekday from df but you could recompute)
            month_df["date_str"] = month_df["date"].dt.strftime("%Y-%m-%d")
            excel_df = pd.DataFrame(
                {
                    "Fecha": month_df["date_str"],
                    "Día de la semana": month_df["weekday"],
                    "Turno": month_df["shift_name"],
                }
            )

            # Safe, valid sheet name
            raw_name = month_df["month_name"].iloc[0]
            sheet_name = (
                raw_name[:31]
                .replace("/", "_")
                .replace("\\", "_")
                .replace("*", "_")
                .replace("?", "_")
                .replace(":", "_")
                .replace("[", "_")
                .replace("]", "_")
            )

            excel_df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            sheets_created += 1

            # Freeze header row
            ws.freeze_panes = "A2"

            # Apply row fills and merge identical dates/weekday cells
            # Determine weekend/holiday from actual date + is_festivo
            # Build a mapping from row -> style first
            row_styles = []
            for _, row in month_df.iterrows():
                is_weekend = row["date"].weekday() >= 5  # 5,6 -> Sat, Sun
                is_holiday = bool(row.get("is_festivo", False))
                if is_holiday:
                    row_styles.append(festivo_fill)
                elif is_weekend:
                    # Use the appropriate weekend fill based on day of week
                    day_of_week = row["date"].weekday()
                    if day_of_week == 5:  # Saturday
                        row_styles.append(saturday_fill)
                    elif day_of_week == 6:  # Sunday
                        row_styles.append(sunday_fill)
                else:
                    # Use specific weekday fill based on day of week
                    day_of_week = row["date"].weekday()
                    row_styles.append(weekday_fills[day_of_week])

            # Color data rows (start at Excel row 2)
            for i, fill in enumerate(row_styles, start=2):
                for col_idx in range(1, 4):  # A..C
                    ws.cell(row=i, column=col_idx).fill = fill

            # Merge consecutive rows that share the same Date (and Weekday) so Date/Weekday spans multiple shifts
            # We'll iterate groups by the *displayed* date string
            data_len = len(excel_df)
            if data_len > 0:
                start = 2
                prev_date = excel_df.iloc[0]["Fecha"]
                prev_weekday = excel_df.iloc[0]["Día de la semana"]

                for r in range(3, data_len + 2 + 1):  # next data row through one-past-the-end sentinel
                    if r <= data_len + 1:
                        this_date = excel_df.iloc[r - 2]["Fecha"]
                        this_weekday = excel_df.iloc[r - 2]["Día de la semana"]
                    else:
                        this_date = None
                        this_weekday = None

                    if this_date != prev_date or this_weekday != prev_weekday:
                        end = r - 1
                        if end > start:  # only merge if more than one row
                            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)  # A
                            ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)  # B
                            ws.cell(row=start, column=1).alignment = header_align
                            ws.cell(row=start, column=2).alignment = header_align
                        # reset window
                        start = r
                        prev_date = this_date
                        prev_weekday = this_weekday

            # Auto-fit columns (cap width to 50)
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    try:
                        v = "" if cell.value is None else str(cell.value)
                        max_len = max(max_len, len(v))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

            # Row height for readability
            for r in range(2, data_len + 2):
                ws.row_dimensions[r].height = 25

    if sheets_created == 0:
        return None

    excel_data.seek(0)
    return excel_data.getvalue()

# excel_builder.py
from io import BytesIO
import pandas as pd
import numpy as np

def generate_monthly_assignments_excel_bytes(df: pd.DataFrame) -> bytes:
    """
    Same behavior as before, but returns the .xlsx file as bytes
    for use in Streamlit's download button.
    """
    required = {"date", "day_of_week", "section_name", "worker_name", "libra", "is_festivo"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"DataFrame missing required columns: {sorted(missing)}")

    df = df.copy()
    df["date"] = pd.to_datetime(df["date"]).dt.normalize()
    df["section_name"] = df["section_name"].astype(str)
    df["worker_name"] = df["worker_name"].astype(str)
    df["libra"] = df["libra"].astype(bool)
    df["is_festivo"] = df["is_festivo"].astype(bool)
    df["worker_display"] = np.where(df["libra"], df["worker_name"] + " (L)", df["worker_name"])
    df["month_key"] = df["date"].dt.to_period("M").astype(str)

    day_flags = (
        df[["date", "is_festivo"]]
        .drop_duplicates()
        .assign(weekday_num=lambda x: x["date"].dt.weekday)
        .assign(is_weekend=lambda x: x["weekday_num"] >= 5)
    )

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter", datetime_format="yyyy-mm-dd") as writer:
        wb = writer.book
        fmt_header = wb.add_format({"bold": True, "align": "center", "valign": "vcenter"})
        fmt_header_date = wb.add_format({"bold": True, "align": "center", "valign": "vcenter", "num_format": "dd-mmm"})
        fmt_header_dow = wb.add_format({"align": "center", "valign": "vcenter", "font_color": "#555555"})
        fmt_weekend = wb.add_format({"bg_color": "#FFF8E1"})
        fmt_festivo = wb.add_format({"bg_color": "#F8D7DA"})
        fmt_altrow = wb.add_format({"bg_color": "#F6F8FA"})
        fmt_section = wb.add_format({"bold": True})
        fmt_center = wb.add_format({"align": "center"})
        fmt_wrap = wb.add_format({"text_wrap": True, "valign": "top"})
        fmt_legend = wb.add_format({"font_color": "#555555", "italic": True})
        fmt_border_bottom = wb.add_format({"bottom": 1, "bottom_color": "#DDDDDD"})

        palette = [
            "#E8F5E9", "#E3F2FD", "#FFF8E1", "#F3E5F5", "#E0F7FA", "#FFF3E0", "#EDE7F6",
            "#F1F8E9", "#F9FBE7", "#E8EAF6", "#FCE4EC", "#E0F2F1"
        ]

        for month_key, dfm in df.groupby("month_key"):
            month_first = pd.to_datetime(month_key + "-01")
            month_last = (month_first + pd.offsets.MonthEnd(1)).normalize()
            dates = pd.date_range(month_first, month_last, freq="D")

            flags_month = pd.DataFrame({"date": dates}).merge(day_flags, on="date", how="left")
            flags_month["is_festivo"] = flags_month["is_festivo"].fillna(False)
            flags_month["weekday_num"] = flags_month["date"].dt.weekday
            flags_month["is_weekend"] = flags_month["weekday_num"] >= 5

            joined = (
                dfm.groupby(["section_name", "date"])["worker_display"]
                   .apply(lambda s: ", ".join(sorted(set(s))))
                   .unstack(fill_value="")
                   .reindex(columns=dates, fill_value="")
            )

            ws = wb.add_worksheet(month_key)

            ws.write(0, 0, f"Assignments — {month_key}", fmt_header)

            start_row, start_col = 3, 0
            ws.write(start_row, start_col, "Section", fmt_header)
            for j, d in enumerate(dates, start=start_col + 1):
                ws.write_datetime(start_row, j, d, fmt_header_date)
                ws.write(start_row + 1, j, d.strftime("%a"), fmt_header_dow)
            ws.write(start_row + 1, start_col, "", fmt_header)

            sections = list(joined.index)
            for i, section in enumerate(sections):
                row = start_row + 2 + i
                ws.write(row, start_col, section, wb.add_format({"bold": True, "bg_color": palette[i % len(palette)]}))
                for j, d in enumerate(dates, start=start_col + 1):
                    text = joined.at[section, d] if d in joined.columns else ""
                    cell_fmt = wb.add_format({"text_wrap": True, "valign": "top"})
                    flag_row = flags_month.loc[flags_month["date"] == d]
                    is_weekend = bool(flag_row["is_weekend"].iloc[0]) if not flag_row.empty else False
                    is_festivo = bool(flag_row["is_festivo"].iloc[0]) if not flag_row.empty else False
                    if is_festivo:
                        cell_fmt.set_bg_color("#F8D7DA")
                    elif is_weekend:
                        cell_fmt.set_bg_color("#FFF8E1")
                    elif i % 2 == 1:
                        cell_fmt.set_bg_color("#F6F8FA")
                    ws.write(row, j, text, cell_fmt)

            ws.set_row(start_row + 1, None, fmt_border_bottom)
            ws.freeze_panes(start_row + 2, start_col + 1)

            sec_width = max([len("Section")] + [len(s) for s in sections]) + 2
            ws.set_column(start_col, start_col, min(max(sec_width, 12), 32))
            for j, d in enumerate(dates, start=start_col + 1):
                col_vals = list(joined[d].astype(str).values)
                longest = max([len(d.strftime("%d-%b")), len("Wed")] + [len(v) for v in col_vals]) + 2
                ws.set_column(j, j, min(max(10, longest), 22))
            ws.set_row(start_row, 20)
            ws.set_row(start_row + 1, 16)
    buf.seek(0)
    return buf.getvalue()

def main():
    st.title("📅 Calendari de guàrdies")
    
    # Sidebar with month selector
    st.sidebar.title("Opcions")
    view_option = st.sidebar.radio(
        "Selecciona el mode view:", 
        ["Calendari mensual", "Assignacions","Estadístiques", "Comptatges"]
    )
    
    global month_names
    month_names = ["Gener", "Febrer", "Març", "Abril", "Maig", "Juny", "Juliol", "Agost", "Setembre", "Octubre", "Novembre", "Desembre"]
    year = 2026
    
    # Get data for the selected month
    df = get_shifts_data(year)
    
    if view_option == "Calendari mensual":
        selected_month = st.selectbox("Selecciona el mes:", month_names)
        selected_month_num = month_names.index(selected_month) + 1
        
        year = 2026
        # Filter data for the selected month
        month_df = df[df["month"] == selected_month_num]
        
        # Existing calendar view code...
        st.header(f"{selected_month} {year}")
        draw_month_calendar(month_df, selected_month, year)

        # Generate Excel file when user clicks the download button
        excel_data = create_calendar_excel(get_shifts_data(year))


        st.download_button(
            label="⬇️ Descarrega el calendari anual en Excel",
            data=excel_data,
            file_name=f"calendario_{year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


    elif view_option == "Assignacions":
        selected_month = st.selectbox("Selecciona el mes:", month_names)
        selected_month_num = month_names.index(selected_month) + 1
        
        year = 2026
        try:
            # Load assignments
            assignacions = db.get_assignment_scenarios()
            if isinstance(assignacions, pd.DataFrame) and not assignacions.empty:
                # Create identifiers using DataFrame column access
                nuevas_ids = {row['id'] : f"{row.created_by} - {row.description} - {(row.created_at).split('T')[0]}" for _, row in assignacions.iterrows()}

                choose = st.selectbox("Selecciona les assignacions a visualitzar:", nuevas_ids.values())

                selected_id = [key for key, val in nuevas_ids.items() if val == choose][0]
                assignments_df = db.get_assignments(selected_id)
                # Convert date column to datetime
                assignments_df['date'] = pd.to_datetime(assignments_df['date'])

                # Filter the data
                filtered_df = assignments_df.copy()

                # Get all shifts data for the selected month to show in calendar
                shifts_df = get_shifts_data(year, selected_month_num)
                st.subheader(f"{selected_month} {year}")
                # Draw calendar with assignments
                month_filtered_df = filtered_df[filtered_df['date'].dt.month == selected_month_num]
                if month_filtered_df.empty:
                    st.warning(f"No s'han trobat assignacions pel mes {selected_month} {year}. Mostrant calendari base")
                    draw_month_calendar(shifts_df, selected_month, year)

                draw_assignment_calendar(shifts_df, month_filtered_df, selected_month, year)
                def build_bytes(_df: pd.DataFrame) -> bytes:
                    return generate_monthly_assignments_excel_bytes(_df)

                xlsx_bytes = build_bytes(filtered_df)

                st.download_button(
                    label="📥 Descarrega les assignacions en Excel",
                    data=xlsx_bytes,
                    file_name="Monthly_Assignments.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            
            else:
                st.warning("No s'han trobat assignacions")

        except Exception as e:
            st.error(f"Error loading assignments: {str(e)}")

    elif view_option == "Estadístiques":
        selected_month = st.selectbox("Selecciona el mes:", month_names)
        selected_month_num = month_names.index(selected_month) + 1
        st.header(f"Estadístiques per {selected_month} {year}")
        df = df[df["month"] == selected_month_num]
        year = 2026
        if not df.empty:
            # Create tabs for different statistics views
            stats_tabs = st.tabs(["Resum Mensual", "Distribució de Torns", "Comparació de Treballadors"])
            
            with stats_tabs[0]:
                # Basic statistics for the selected month
                total_shifts = len(df)
                total_hours = df['hours'].sum()
                total_personnel_needed = df['personnel'].sum()
                night_shifts = df[df['shift_name'].str.contains('noc', case=False)].shape[0]
                weekend_shifts = df[df['weekday'].isin(['saturday', 'sunday', 'festivo'])].shape[0]
                festivo_shifts = df[df['is_festivo'] == True].shape[0]
                
                # Mostra les mètriques en columnes
                col1, col2, col3 = st.columns(3)
                col1.metric("Torns Totals", total_shifts)
                col2.metric("Hores Totals", f"{total_hours:,.1f}")
                col3.metric("Personal Necessari", total_personnel_needed)
                
                col1, col2, col3 = st.columns(3)
                col1.metric("Torns de Nit", night_shifts)
                col2.metric("Torns de Cap de Setmana", weekend_shifts)
                col3.metric("Torns de Festiu", festivo_shifts)
                
            with stats_tabs[1]:
                # Distribució de torns per tipus
                st.subheader("Distribució de Torns per Tipus")
                shift_counts = df['shift_name'].value_counts().reset_index()
                shift_counts.columns = ['Tipus de Torn', 'Comptatge']
                
                # Crear un gràfic de barres per tipus de torns
                shift_chart = alt.Chart(shift_counts).mark_bar().encode(
                    y=alt.Y('Tipus de Torn:N', sort='-y', title=None, axis=alt.Axis(labelLimit=350)),
                    x='Comptatge:Q',
                    color=alt.Color('Tipus de Torn:N', 
                                  scale=alt.Scale(range=['#4285F4', '#34A853', '#FBBC05', '#EA4335', '#9AA0A6']), legend=None),
                    tooltip=['Tipus de Torn', 'Comptatge']
                ).properties(
                    width=600,
                    height=400
                )
                
                st.altair_chart(shift_chart, use_container_width=False)
            with stats_tabs[2]:
                st.subheader("Comparació per seccions i treballadors")
                
                # Try to load assignment data
                try:
                    # Try to load assignment data from the database instead of a static file
                        # Load available assignment scenarios
                    assignacions = db.get_assignment_scenarios()
                    
                    if isinstance(assignacions, pd.DataFrame) and not assignacions.empty:
                        # Create identifiers using DataFrame column access
                        scenario_ids = {row['id']: f"{row.created_by} - {row.description} - {(row.created_at).split('T')[0]}" 
                                        for _, row in assignacions.iterrows()}
                        
                        # Let user select which assignment scenario to analyze
                        chosen_scenario = st.selectbox(
                            "Selecciona l'assignació per analitzar:", 
                            options=list(scenario_ids.values()),
                            key="stats_scenario_selector"
                        )
                        
                        # Get the ID of the selected scenario
                        selected_id = [key for key, val in scenario_ids.items() if val == chosen_scenario][0]
                        
                        # Load the assignments for the selected scenario
                        assignments_df = db.get_assignments(selected_id)
                        
                        # Convert date column to datetime
                        assignments_df['date'] = pd.to_datetime(assignments_df['date'])
                        
                        # Filter for the selected month
                        month_assignments = assignments_df[
                            (assignments_df['date'].dt.month == selected_month_num) & 
                            (assignments_df['date'].dt.year == year)
                        ]
                        
                        if month_assignments.empty:
                            st.warning(f"No s'han trobat assignacions per {selected_month} {year} en aquest escenari.")
                        else:
                            # Agrupar per treballador i secció per obtenir els comptatges
                            worker_section_counts = month_assignments.groupby(['worker_name', 'section_name']).size().reset_index(name='shift_count')
                            
                            # Crear una taula pivot per a una millor visualització
                            pivot_df = worker_section_counts.pivot(
                                index='worker_name',
                                columns='section_name',
                                values='shift_count'
                            ).fillna(0).astype(int)
                            
                            # Afegir una columna de total
                            pivot_df['Total'] = pivot_df.sum(axis=1)
                            
                            # Ordenar per treballador
                            pivot_df = pivot_df.sort_values('worker_name', ascending=True)
                            # Rename worker_name to Treballador in worker_section_counts
                            worker_section_counts = worker_section_counts.rename(columns={'worker_name': 'Treballador'})

                            # Create pivot table with the renamed column
                            pivot_df = worker_section_counts.pivot(
                                index='Treballador',
                                columns='section_name',
                                values='shift_count'
                            ).fillna(0).astype(int)

                            # Add a total column
                            pivot_df['Total'] = pivot_df.sum(axis=1)

                            # Add areas column to pivot_df
                            # Get workers from the database
                            workers = db.get_workers()
                            uci = {}
                            urg = {}
                            for worker in workers:
                                # Extract areas information if available
                                if hasattr(worker, 'areas') and worker.areas:
                                    uci[worker.name] = "Sí" if "Guardia_UCI" in worker.areas else "No"
                                    urg[worker.name] = "Sí" if "Guardia_Urg" in worker.areas else "No"
                                    areas_str = ", ".join(worker.areas)
                                else:
                                    uci[worker.name] = "No"
                                    urg[worker.name] = "No"
                                # Add the areas column to the pivot_df
                            pivot_df['UCI'] = pd.Series(uci)
                            pivot_df['Urg'] = pd.Series(urg)

                            # Reorder columns to make 'Àrees' the second column
                            cols = list(pivot_df.columns)
                            cols.remove('UCI')
                            cols.remove('Urg')  
                            cols.insert(0, 'UCI')
                            cols.insert(1, 'Urg')
                            pivot_df = pivot_df[cols]
                            # Add the areas column to the pivot_df
                            # Mostrar la taula pivot
                            st.write("##### Torns per Treballador i Secció")
                            st.dataframe(pivot_df.style.background_gradient(cmap='Blues', axis=None))
                            
                            # Crear una taula resum amb percentatges
                            st.write("##### Distribució d'Assignacions per Treballador")
                            
                            # Obtenir el total de torns per secció
                            section_totals = worker_section_counts.groupby('section_name')['shift_count'].sum().reset_index()
                            section_totals.columns = ['section_name', 'total_shifts']
                            
                            # Fusionar per calcular percentatges
                            worker_section_pct = worker_section_counts.merge(section_totals, on='section_name')
                            worker_section_pct['percentage'] = (worker_section_pct['shift_count'] / worker_section_pct['total_shifts'] * 100).round(1)
                            
                            # Mostrar la distribució de percentatges
                            section_pivot = worker_section_pct.pivot(
                                index='section_name',
                                columns='Treballador',
                                values='percentage'
                            ).fillna(0)
                            
                            st.dataframe(section_pivot.style.background_gradient(cmap='Greens', axis=None).format("{:.1f}%"))
                            
                            # Calcular mitjanes mensuals per secció i treballador
                            st.write("##### Mitjana Mensual de Torns per Secció i Treballador")
                            
                            # Comptar els dies del mes
                            days_in_month = calendar.monthrange(year, selected_month_num)[1]
                            
                            # Obtenir treballadors i seccions únics
                            workers = month_assignments['worker_name'].unique()
                            sections = month_assignments['section_name'].unique()
                            
                            # Crear un dataframe per a les mitjanes diàries
                            monthly_means = []
                            
                            for worker in workers:
                                worker_data = {
                                    'Treballador': worker,
                                    'Torns Totals': len(month_assignments[month_assignments['worker_name'] == worker]),
                                    'Mitjana Diària': len(month_assignments[month_assignments['worker_name'] == worker]) / days_in_month
                                }
                                
                                # Afegir mitjanes específiques per secció
                                for section in sections:
                                    section_count = len(month_assignments[
                                        (month_assignments['worker_name'] == worker) & 
                                        (month_assignments['section_name'] == section)
                                    ])
                                    worker_data[f"{section} (mitjana/dia)"] = section_count / days_in_month
                                
                                monthly_means.append(worker_data)
                            
                            monthly_means_df = pd.DataFrame(monthly_means)
                            monthly_means_df = monthly_means_df.sort_values('Torns Totals', ascending=False)
                            # Crear gràfics de comparació
                            st.write("##### Comparació Visual d'Assignacions per Treballador")
                            
                            # 1. Gràfic de barres dels torns totals per treballador
                            chart_data = monthly_means_df[['Treballador', 'Torns Totals']].sort_values('Torns Totals', ascending=False)
                            
                            # Sort data by Torns Totals (descending) and then by Treballador for ties
                            chart_data = chart_data.sort_values(['Torns Totals', 'Treballador'], ascending=[False, True])

                            # Create bar chart with pre-sorted data (no need to specify sort in encoding)
                            total_shifts = month_assignments.shape[0]
                            worker_load = month_assignments['worker_name'].value_counts().reset_index()
                            worker_load.columns = ['Treballador', 'Torns']
                            worker_load['Percentatge'] = (worker_load['Torns'] / total_shifts * 100).round(1)
                            worker_load = worker_load.sort_values(['Torns', 'Treballador'], ascending=[False, True])

                            balance_chart = alt.Chart(worker_load).mark_bar().encode(
                                x=alt.X('Treballador:N', sort='-y', title='Treballador', axis=alt.Axis(labelAngle=45)),
                                y=alt.Y('Percentatge:Q', title='% de Torns Totals'),
                                color=alt.Color('Treballador:N'),
                                tooltip=['Treballador', 'Torns', 'Percentatge']
                            ).properties(
                                width=600,
                                height=300,
                                title="Distribució de Càrrega de Treball (% de Torns Totals)"
                            )

                            ideal = alt.Chart(pd.DataFrame({'y': [100 / len(workers)]})).mark_rule(
                                color='red', 
                                strokeDash=[3, 3]
                            ).encode(y='y:Q')
                            
                            st.altair_chart(balance_chart + ideal, use_container_width=True)
                            
                            worker_section_counts = worker_section_counts.sort_values(['shift_count', 'Treballador'], ascending=[False, True])

                            # 3. Distribució de torns per treballador i secció (barres apilades)
                            stacked_chart = alt.Chart(worker_section_counts).mark_bar().encode(
                                x=alt.X('Treballador:N', title='Treballador', sort='-y', axis=alt.Axis(labelAngle=45)),
                                y=alt.Y('shift_count:Q', title='Nombre de Torns'),
                                color=alt.Color('section_name:N', title='Secció'),
                                tooltip=['Treballador', 'section_name', 'shift_count']
                            ).properties(
                                width=600,
                                height=400,
                                title="Distribució de Torns per Treballador i Secció"
                            )
                            
                            st.altair_chart(stacked_chart, use_container_width=True)
                            
                    else:
                        st.warning("No s'han trobat escenaris d'assignació en la base de dades.")
                        month_assignments = pd.DataFrame()  # Empty DataFrame as fallback
                except Exception as e:
                    st.error(f"Error carregant les assignacions: {str(e)}")

    elif view_option == "Comptatges":
        # In the main() function, change the view_option radio to include "Contaje"
        st.header("Comptatge per Treballador")
        try:
            # Try to load assignment data from the database instead of a static file
                # Load available assignment scenarios
            assignacions = db.get_assignment_scenarios()
            
            if isinstance(assignacions, pd.DataFrame) and not assignacions.empty:
                # Create identifiers using DataFrame column access
                scenario_ids = {row['id']: f"{row.created_by} - {row.description} - {(row.created_at).split('T')[0]}" 
                                for _, row in assignacions.iterrows()}
                
                # Let user select which assignment scenario to analyze
                chosen_scenario = st.selectbox(
                    "Selecciona l'assignació per analitzar:", 
                    options=list(scenario_ids.values()),
                    key="stats_scenario_selector"
                )
                
                # Get the ID of the selected scenario
                selected_id = [key for key, val in scenario_ids.items() if val == chosen_scenario][0]
                
                # Load the assignments for the selected scenario
                assignments_df = db.get_assignments(selected_id)
                
                # Convert date column to datetime
                assignments_df['date'] = pd.to_datetime(assignments_df['date'])
            
                col1, col2 = st.columns(2)
                with col1:
                    start_month = st.selectbox(
                        "Mes inicial",
                        options=list(range(1, 13)),
                        index=0,
                        format_func=lambda x: [
                        "Gener", "Febrer", "Març", "Abril", "Maig", "Juny",
                        "Juliol", "Agost", "Setembre", "Octubre", "Novembre", "Desembre"
                        ][x - 1])
                with col2:
                    end_month = st.selectbox(
                    "Mes final",
                    options=list(range(1, 13)),
                    index=11,
                    format_func=lambda x: [
                    "Gener", "Febrer", "Març", "Abril", "Maig", "Juny",
                    "Juliol", "Agost", "Setembre", "Octubre", "Novembre", "Desembre"
                    ][x - 1]
                )

                if start_month <= end_month:
                    # Regular range (e.g., April to August)
                    assignments_df = assignments_df[(assignments_df['date'].dt.month >= start_month) & 
                                                (assignments_df['date'].dt.month <= end_month)]
                else:
                    # Range spans across year boundary (e.g., November to February)
                    assignments_df = assignments_df[(assignments_df['date'].dt.month >= start_month) | 
                                                (assignments_df['date'].dt.month <= end_month)]
                
                if not assignments_df.empty:
                    # Get sections from the database (ensure it's available in local scope)
                    local_sections = db.get_sections()
                    
                    # Calculate hours based on the new rules
                    def calculate_hours(row, sections_list, assignments_data):
                        shift_type = row['section_name']
                        weekday = row['date'].weekday()  # 0=Monday, 4=Friday, 5=Saturday, 6=Sunday
                        is_festivo = row['is_festivo']
                        is_weekend = row['is_weekend']
                        date = row['date']
                        
                        # Initialize the counters
                        total_hours = 0
                        
                        # UCI_G_lab: always counts 7 hours
                        if shift_type == 'UCI_G_lab':
                            total_hours = 7
                        
                        # UCI_G_festivo: counts 24 hours, except when there's also HEMS_festivo that day (then 16)
                        elif shift_type == 'UCI_G_festivo':
                            # Check if there's a HEMS_festivo on the same day
                            same_day_assignments = assignments_data[assignments_data['date'].dt.date == date.date()]
                            has_hems_festivo = any('HEMS_festivo' in section for section in same_day_assignments['section_name'])
                            
                            if has_hems_festivo:
                                total_hours = 16
                            else:
                                total_hours = 24
                        
                        # HEMS_festivo: always counts 16 hours
                        elif shift_type == 'HEMS_festivo':
                            total_hours = 16
                        
                        # Coordis_diurno: counts 12 hours on Saturday and Sunday, 0 the rest
                        elif shift_type == 'Coordis_diurno':
                            if weekday in [5, 6]:  # Saturday, Sunday
                                total_hours = 12
                            else:
                                total_hours = 0
                        
                        # Coordis_nocturno: counts 12 hours on Friday, Saturday and Sunday, 0 the rest
                        elif shift_type == 'Coordis_nocturno':
                            if weekday in [4, 5, 6]:  # Friday, Saturday, Sunday
                                total_hours = 12
                            else:
                                total_hours = 0
                        
                        # For all Urg shifts, use the hours from the section's horas parameter
                        elif 'Urg_G' in shift_type:
                            # Find the section object and get its horas_turno
                            section_obj = next((s for s in sections_list if s.nombre == shift_type), None)
                            if section_obj:
                                total_hours = section_obj.horas_turno
                            else:
                                total_hours = 0  # Default if section not found
                        
                        # For any other shifts, try to get hours from section
                        else:
                            section_obj = next((s for s in sections_list if s.nombre == shift_type), None)
                            if section_obj:
                                total_hours = section_obj.horas_turno
                            else:
                                total_hours = 0  # Default if section not found
                        
                        return pd.Series([total_hours])
                    
                    # Apply calculation to each row with local_sections and assignments_df as arguments
                    assignments_df[['total_hours']] = assignments_df.apply(
                        lambda row: calculate_hours(row, local_sections, assignments_df), axis=1
                    )
                    
                    # Create tabs for different views
                    tab1, tab2 = st.tabs(["Resum per Treballador", "Detall de Torns"])
                    def torns_per_worker(assignments_df):
                        month_assignments = assignments_df.copy()
                        if not month_assignments.empty:
                            # Agrupar per treballador i secció per obtenir els comptatges
                            worker_section_counts = month_assignments.groupby(['worker_name', 'section_name']).size().reset_index(name='shift_count')
                            
                            # Crear una taula pivot per a una millor visualització
                            pivot_df = worker_section_counts.pivot(
                                index='worker_name',
                                columns='section_name',
                                values='shift_count'
                            ).fillna(0).astype(int)
                            
                            # Afegir una columna de total
                            pivot_df['Total'] = pivot_df.sum(axis=1)
                            
                            # Ordenar per treballador
                            pivot_df = pivot_df.sort_values('worker_name', ascending=True)
                            # Rename worker_name to Treballador in worker_section_counts
                            worker_section_counts = worker_section_counts.rename(columns={'worker_name': 'Treballador'})

                            # Create pivot table with the renamed column
                            pivot_df = worker_section_counts.pivot(
                                index='Treballador',
                                columns='section_name',
                                values='shift_count'
                            ).fillna(0).astype(int)

                            # Add a total column
                            pivot_df['Total'] = pivot_df.sum(axis=1)

                            # Add areas column to pivot_df
                            # Get workers from the database
                            workers = db.get_workers()
                            uci = {}
                            urg = {}
                            for worker in workers:
                                # Extract areas information if available
                                if hasattr(worker, 'areas') and worker.areas:
                                    uci[worker.name] = "Sí" if "Guardia_UCI" in worker.areas else "No"
                                    urg[worker.name] = "Sí" if "Guardia_Urg" in worker.areas else "No"
                                    areas_str = ", ".join(worker.areas)
                                else:
                                    uci[worker.name] = "No"
                                    urg[worker.name] = "No"
                                # Add the areas column to the pivot_df
                            pivot_df['UCI'] = pd.Series(uci)
                            pivot_df['Urg'] = pd.Series(urg)

                            # Reorder columns to make 'Àrees' the second column
                            cols = list(pivot_df.columns)
                            cols.remove('UCI')
                            cols.remove('Urg')  
                            cols.insert(0, 'UCI')
                            cols.insert(1, 'Urg')
                            pivot_df = pivot_df[cols]
                            # Add the areas column to the pivot_df
                            # Mostrar la taula pivot
                            st.write("##### Torns per Treballador i Secció")
                            st.dataframe(pivot_df.style.background_gradient(cmap='Blues', axis=None))

                    with tab1:
                        torns_per_worker(assignments_df)
                        # Group by worker and calculate totals
                        worker_summary = assignments_df.groupby('worker_name').agg({
                            'total_hours': 'sum',
                            'date': 'count'
                        }).reset_index()
                        
                        worker_summary.columns = ['Treballador', 'Hores Totals', 'Total de Torns']
                        
                        # Add UCI and Urg information
                        # Get workers from the database
                        workers = db.get_workers()
                        uci = {}
                        urg = {}
                        for worker in workers:
                            # Extract areas information if available
                            if hasattr(worker, 'areas') and worker.areas:
                                uci[worker.name] = "Sí" if "Guardia_UCI" in worker.areas else "No"
                                urg[worker.name] = "Sí" if "Guardia_Urg" in worker.areas else "No"
                            else:
                                uci[worker.name] = "No"
                                urg[worker.name] = "No"
                        
                        # Add the areas columns to the worker_summary
                        worker_summary['UCI'] = worker_summary['Treballador'].map(uci)
                        worker_summary['Urg'] = worker_summary['Treballador'].map(urg)
                        
                        # Reorder columns to put UCI and Urg first
                        cols = ['Treballador', 'UCI', 'Urg', 'Hores Totals', 'Total de Torns']
                        worker_summary = worker_summary[cols]
                        
                        # Display summary in tabs by worker type
                        st.subheader("Resum d'Hores per Treballador")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            st.write("Treballadors d'UCI")
                            # Filter for UCI workers
                            uci_workers = worker_summary[worker_summary['UCI'] == 'Sí'].copy()
                            if not uci_workers.empty:
                                st.dataframe(uci_workers.style.format({'Hores Totals': '{:.0f}'}).background_gradient(cmap='Blues', subset=['Hores Totals']))
                            else:
                                st.info("No s'han trobat treballadors d'UCI en aquest període.")

                        with col2:
                            st.write("Treballadors d'Urgències")
                            # Filter for Urg workers
                            urg_workers = worker_summary[worker_summary['Urg'] == 'Sí'].copy()
                            if not urg_workers.empty:
                                st.dataframe(urg_workers.style.format({'Hores Totals': '{:.0f}'}).background_gradient(cmap='Reds', subset=['Hores Totals']))
                            else:
                                st.info("No s'han trobat treballadors d'Urgències en aquest període.")
                        
                        # def to_excel_with_formatting(df):
                        #     output = io.BytesIO()
                        #     with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        #         df.to_excel(writer, sheet_name='Resum', index=False)
                        #         workbook = writer.book
                        #         worksheet = writer.sheets['Resum']
                        
                        #         # Get range for gradient formatting
                        #         max_row = len(df) + 1
                        #         col_index = df.columns.get_loc('Hores Totals')
                        
                        #         # Apply conditional formatting to 'Hores Totals' column (0-based indexing)
                        #         worksheet.conditional_format(1, col_index, max_row, col_index, {
                        #             'type': '3_color_scale',
                        #             'min_color': "#ffffff",
                        #             'mid_color': "#ff9999",
                        #             'max_color': "#990000"
                        #         })
                        
                        #     output.seek(0)
                        #     return output.getvalue()  # Return the binary data, not the BytesIO object
                        # # Download button
                        
                        #                     # In tab1, replace this section:
                        # excel_data = to_excel_with_formatting(worker_summary)
                        # st.download_button(
                        #     label="📥 Descarrega l'Excel",
                        #     data=excel_data,
                        #     file_name='resum_treballadors.xlsx',
                        #     mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        # )
                        
                    with tab2:
                        # Group by worker and shift type
                        shift_breakdown = assignments_df.groupby(['worker_name', 'section_name']).agg({
                            'total_hours': 'sum',
                            'date': 'count'  # Count dates instead of section_name
                        }).reset_index()
                        
                        shift_breakdown.columns = ['Treballador', 'Tipus de Torn', 'Hores Totals', 'Número de Torns']
                        
                        # Sort by worker and total hours
                        shift_breakdown = shift_breakdown.sort_values(['Treballador', 'Hores Totals'], ascending=[True, False])
                        
                        # Display breakdown
                        st.subheader("Detall de Torns per Treballador")
                        
                        st.dataframe(shift_breakdown)
                        
                        # Create pivot table
                        pivot_df = shift_breakdown.pivot_table(
                            index='Treballador',
                            columns='Tipus de Torn',
                            values='Hores Totals',
                            aggfunc='sum'
                        ).fillna(0)
                        
                        # Add a total column
                        pivot_df['Total'] = pivot_df.sum(axis=1)
                        
                        # Display pivot table with 1 decimal formatting
                        st.subheader("Taula Pivot d'Hores per Tipus de Torn")
                        st.dataframe(pivot_df.style.format("{:.0f}").background_gradient(cmap='YlGnBu'))
        except Exception as e:
            st.error(f"Error carregant les assignacions: {str(e)}")
            import traceback
            st.code(traceback.format_exc())

if __name__ == "__main__":
    main()